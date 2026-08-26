#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
run_match.py —— JD 批量匹配 CLI（零 LLM 调用）

用法：
    python run_match.py \\
        --profile candidate_profile.json \\
        --config match_config.yaml \\
        --input jds.csv \\
        [--out result.csv] [--limit 20] [--embed]

输入：CSV（utf-8-sig），至少含 岗位名称/公司名称/薪资范围/工作地点/JD描述 等列。
输出：源文件全部列 + 追加 [优先级判断, 匹配度, 优先级判断理由, 能力条目得分, 深度匹配]，
      保存为 <原名>_判定结果.csv（若路径含 01_raw/ 则写到同级 02_result/）。

依赖：PyYAML（必需）；openpyxl（仅当输入/输出为 .xlsx 时需要）；
      sentence-transformers+faiss-cpu（仅 --embed 时需要）。
"""

import argparse
import csv
import sys
import yaml
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from matcher import load_profile, load_config, Matcher, enrich_jd

APPEND_COLS = ["优先级判断", "匹配度", "优先级判断理由", "能力条目得分", "叠加信号", "平均JD深度", "平均候选人深度"]


def read_rows(path):
    if path.lower().endswith(".xlsx"):
        try:
            import openpyxl
        except ImportError:
            sys.stderr.write("读取 xlsx 需要 openpyxl：pip install openpyxl\n")
            sys.exit(2)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(h) if h is not None else "" for h in rows[0]]
        return [dict(zip(header, ["" if c is None else c for c in r])) for r in rows[1:]]
    # CSV：必须用 DictReader 统计，禁止 wc -l（JD 描述含换行）
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return [dict(r) for r in reader]


def write_rows(path, header, rows):
    if path.lower().endswith(".xlsx"):
        try:
            import openpyxl
        except ImportError:
            sys.stderr.write("写出 xlsx 需要 openpyxl：pip install openpyxl\n")
            sys.exit(2)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(header)
        for r in rows:
            ws.append([r.get(h, "") for h in header])
        wb.save(path)
        return
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=header)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def dedup(rows):
    seen = set()
    out = []
    dup = 0
    for r in rows:
        key = (r.get("job_id") or "").strip() or (
            str(r.get("岗位名称", "")) + "|" + str(r.get("公司名称", ""))
        )
        if key in seen:
            dup += 1
            continue
        seen.add(key)
        out.append(r)
    return out, dup


def decide_out_path(in_path):
    p = Path(in_path)
    if "01_raw" in p.parts:
        idx = p.parts.index("01_raw")
        base = Path(*p.parts[:idx])
        out_dir = base / "02_result"
        out_dir.mkdir(parents=True, exist_ok=True)
        return str(out_dir / f"{p.stem}_判定结果{p.suffix}")
    return str(p.parent / f"{p.stem}_判定结果{p.suffix}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--profile", required=True)
    ap.add_argument("--config", required=True)
    ap.add_argument("--input", required=True)
    ap.add_argument("--out", default=None)
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--embed", action="store_true")
    ap.add_argument("--light-embed", action="store_true", help="启用轻量级语义匹配器（TF-IDF+FAISS，无需torch）")
    ap.add_argument("--api-embed", action="store_true", help="启用 SiliconFlow API embedding（bge-large-zh-v1.5）")
    ap.add_argument("--api-key", default=None, help="SiliconFlow API Key（或设环境变量 SILICONFLOW_API_KEY）")
    args = ap.parse_args()

    profile = load_profile(args.profile)
    config = load_config(args.config)
    embed = None
    light_embed = None
    api_embed = None
    if args.api_embed:
        try:
            from semantic_match_api import APISemanticMatcher
            import os
            tx_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "config/taxonomy.yaml")
            # 优先级: 命令行参数 > 环境变量 > config.local.yaml
            api_key = args.api_key or os.environ.get("SILICONFLOW_API_KEY", "")
            if not api_key:
                local_cfg_path = Path(__file__).parent.parent / "config/config.local.yaml"
                if local_cfg_path.exists():
                    with open(local_cfg_path, encoding="utf-8") as f:
                        local_cfg = yaml.safe_load(f) or {}
                    api_key = local_cfg.get("siliconflow", {}).get("api_key", "")
            if not api_key:
                print("[api-embed] 需要 API Key: --api-key sk-xxx 或 SILICONFLOW_API_KEY 环境变量或 config.local.yaml")
            else:
                api_embed = APISemanticMatcher(tx_path, api_key=api_key)
                api_embed.ensure_built()
                print("[api-embed] SiliconFlow embedding 就绪 (bge-large-zh-v1.5)")
        except Exception as e:
            print(f"[api-embed] 加载失败，回退关键词：{e}")
    elif args.light_embed:
        try:
            from semantic_match_light import LightSemanticMatcher
            import os
            tx_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "config/taxonomy.yaml")
            light_embed = LightSemanticMatcher(tx_path)
            light_embed.ensure_built()
            print("[light-embed] 轻量级语义匹配器就绪 (TF-IDF+FAISS)")
        except Exception as e:
            print(f"[light-embed] 加载失败，回退关键词：{e}")
    elif args.embed:
        try:
            from embed_classify import LocalEmbeddingClassifier
            embed = LocalEmbeddingClassifier(config)
            print("[embed] 本地 embedding 分类器就绪" if embed.is_available
                  else "[embed] 不可用，回退关键词")
        except Exception as e:
            print(f"[embed] 回退关键词：{e}")
    matcher = Matcher(profile, config, embed=embed, light_embed=light_embed, api_embed=api_embed)

    rows = read_rows(args.input)
    rows, dup = dedup(rows)
    print(f"读取 {len(rows)} 条（去重 {dup} 条）")

    header = None
    if args.input.lower().endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(args.input, read_only=True)
        header = [str(h) for h in next(wb.active.iter_rows(min_row=1, max_row=1))]
        wb.close()
    else:
        with open(args.input, "r", encoding="utf-8-sig", newline="") as f:
            header = list(csv.DictReader(f).fieldnames)
    out_header = header + APPEND_COLS

    results = []
    for i, row in enumerate(rows):
        if args.limit and i >= args.limit:
            break
        jd = enrich_jd(row)
        res = matcher.evaluate(jd)
        out_row = dict(row)
        for col in APPEND_COLS:
            out_row[col] = res.get(col, "")
        results.append(out_row)

    out_path = args.out or decide_out_path(args.input)
    write_rows(out_path, out_header, results)

    # 汇报
    from collections import Counter
    cnt = Counter(r["优先级判断"] for r in results)
    total = len(results)
    print("匹配完成")
    print(f"总计：{total} 条（去重后）")
    for lv in ("high", "medium", "low"):
        n = cnt.get(lv, 0)
        pct = f"{n/total*100:.0f}%" if total else "0%"
        print(f"{lv}：{n} 条（{pct}）")
    print(f"结果文件：{out_path}")


if __name__ == "__main__":
    main()
